from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Palette (hex without #) ───────────────────────────────────────────────────
AMBER      = 'D4900A'
AMBER_BG   = 'FFF8E6'
DARK       = '1E293B'
MUTED      = '64748B'
LIGHT_BG   = 'F8FAFC'
BORDER_HEX = 'E2E8F0'
RED_HEX    = 'EF4444'
RED_BG     = 'FEF2F2'
AMBER_RISK = 'F59E0B'
AMBER_RISK_BG = 'FFFBEB'
GREEN_HEX  = '10B981'
GREEN_BG   = 'ECFDF5'
BLUE_HEX   = '3B82F6'
BLUE_BG    = 'EFF6FF'
WHITE      = 'FFFFFF'
BLACK      = '000000'


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)

def _border() -> Border:
    s = Side(border_style='thin', color=BORDER_HEX)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size: int = 12) -> Font:
    return Font(bold=True, color=WHITE, name='Courier New', size=size)

def _label_font() -> Font:
    return Font(bold=True, color=DARK, name='Courier New', size=9)

def _value_font() -> Font:
    return Font(color=DARK, name='Calibri', size=10)

def _muted_font() -> Font:
    return Font(color=MUTED, name='Calibri', size=9)

def _amber_label_font() -> Font:
    return Font(bold=True, color=AMBER, name='Courier New', size=9)

def _score_font(score: int) -> Font:
    if score >= 70: return Font(bold=True, color=RED_HEX,    name='Courier New', size=12)
    if score >= 40: return Font(bold=True, color=AMBER_RISK,  name='Courier New', size=12)
    return          Font(bold=True, color=GREEN_HEX,  name='Courier New', size=12)

def _severity_font(sev: str) -> Font:
    s = sev.lower()
    if s == 'critical': return Font(bold=True, color=RED_HEX,    name='Courier New', size=9)
    if s == 'high':     return Font(bold=True, color=AMBER_RISK,  name='Courier New', size=9)
    if s == 'medium':   return Font(bold=True, color=BLUE_HEX,   name='Courier New', size=9)
    return              Font(bold=True, color=GREEN_HEX,  name='Courier New', size=9)

def _severity_fill(sev: str) -> PatternFill:
    s = sev.lower()
    if s == 'critical': return _fill(RED_BG)
    if s == 'high':     return _fill(AMBER_RISK_BG)
    if s == 'medium':   return _fill(BLUE_BG)
    return _fill(GREEN_BG)

def _col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _sheet_header(ws, title: str, n_cols: int, row: int = 1) -> None:
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f'A{row}:{last_col}{row}')
    c = ws.cell(row, 1)
    c.value     = title
    c.font      = _header_font()
    c.fill      = _fill(AMBER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28

def _section_label(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row, col)
    c.value = text
    c.font  = _amber_label_font()

def _apply_border_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row, col).border = _border()


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel(data: dict) -> bytes:
    trade_dir  = data.get('trade_direction', 'export').upper()
    origin     = data.get('origin', 'Brazil')
    dest       = data.get('destination', 'European Union')
    commodity  = data.get('commodity', 'coffee').capitalize()
    query      = data.get('query', '') or 'Composite trade risk analysis'
    overall    = int(data.get('overall_risk_score', data.get('risk_score', 50)))
    readiness  = int(data.get('export_readiness', data.get('supply_reliability', 100 - overall)))
    reg_score  = int(data.get('regulatory', {}).get('risk_score', 50))
    clim_score = int(data.get('climate',    {}).get('climate_risk_score', 50))
    mkt_score  = int(data.get('market',     {}).get('market_risk_score', 50))
    logi_score = int(data.get('logistics',  {}).get('logistics_risk_score', 50))
    exec_data  = data.get('executive', {})
    exec_sum   = exec_data.get('executive_summary', '') or ''
    verdict    = exec_data.get('overall_verdict', '') or ''
    trade_win  = exec_data.get('trade_window', '') or ''
    key_risks  = exec_data.get('key_risks', []) or []
    actions    = exec_data.get('recommended_actions', []) or []
    gaps       = data.get('gap', {}).get('gaps_identified', []) or []
    origin_port = data.get('logistics', {}).get('origin_port', '') or ''
    dest_port   = data.get('logistics', {}).get('destination_port', '') or ''
    transit     = data.get('logistics', {}).get('estimated_transit_days', 0)
    gps_pct     = data.get('gap', {}).get('supplier_profile', {}).get('gps_coverage_pct', 0)
    now_str     = datetime.now().strftime('%Y-%m-%d %H:%M UTC')

    wb = Workbook()

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1 — Executive Summary
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Executive Summary'
    _col_widths(ws1, [24, 42, 24, 42])

    _sheet_header(ws1, 'ORIGINSIGNAL — Trade Risk Intelligence Report', 4, row=1)

    ws1.merge_cells('A2:D2')
    ws1['A2'].value     = f'Generated: {now_str}'
    ws1['A2'].font      = _muted_font()
    ws1['A2'].alignment = Alignment(horizontal='center')

    # Context
    _section_label(ws1, 4, 1, 'ANALYSIS CONTEXT')
    ws1.merge_cells('A4:D4')

    ctx = [
        (5,  'Commodity',       commodity,  'Trade Direction', trade_dir),
        (6,  'Origin',          origin,     'Destination',     dest),
        (7,  'Query',           query,      '',                ''),
    ]
    for row, k1, v1, k2, v2 in ctx:
        ws1.cell(row, 1).value = k1; ws1.cell(row, 1).font = _label_font()
        ws1.cell(row, 2).value = v1; ws1.cell(row, 2).font = _value_font()
        ws1.cell(row, 2).alignment = Alignment(wrap_text=True)
        ws1.cell(row, 3).value = k2; ws1.cell(row, 3).font = _label_font()
        ws1.cell(row, 4).value = v2; ws1.cell(row, 4).font = _value_font()
        for col in range(1, 5):
            ws1.cell(row, col).fill   = _fill(LIGHT_BG)
            ws1.cell(row, col).border = _border()
    ws1.row_dimensions[7].height = 45

    # Scores
    _section_label(ws1, 9, 1, 'RISK SCORES')
    ws1.merge_cells('A9:D9')

    readiness_label = 'Export Readiness' if trade_dir == 'EXPORT' else 'Supply Reliability'
    score_rows = [
        (10, 'Overall Risk Score', overall),
        (11, 'Regulatory',         reg_score),
        (12, 'Climate',            clim_score),
        (13, 'Market',             mkt_score),
        (14, 'Logistics',          logi_score),
        (15, readiness_label,      readiness),
    ]
    for row, label, score in score_rows:
        bar = '█' * max(1, round(score / 5)) + '░' * (20 - max(1, round(score / 5)))
        ws1.cell(row, 1).value = label;  ws1.cell(row, 1).font = _label_font()
        ws1.cell(row, 2).value = score;  ws1.cell(row, 2).font = _score_font(score)
        ws1.cell(row, 2).alignment = Alignment(horizontal='center')
        ws1.cell(row, 3).value = bar;    ws1.cell(row, 3).font = _score_font(score)
        ws1.merge_cells(f'C{row}:D{row}')
        for col in range(1, 5):
            ws1.cell(row, col).border = _border()

    # Executive Summary text
    _section_label(ws1, 17, 1, 'EXECUTIVE SUMMARY')
    ws1.merge_cells('A17:D17')
    ws1.merge_cells('A18:D18')
    c18 = ws1['A18']
    c18.value     = exec_sum
    c18.font      = _value_font()
    c18.fill      = _fill(AMBER_BG)
    c18.alignment = Alignment(wrap_text=True, vertical='top')
    for col in range(1, 5):
        ws1.cell(18, col).border = _border()
    ws1.row_dimensions[18].height = 90

    # Verdict + trade window
    row = 20
    if verdict:
        ws1.cell(row, 1).value = 'VERDICT';  ws1.cell(row, 1).font = _label_font()
        v_color = GREEN_HEX if verdict == 'Go' else RED_HEX if verdict == 'Hold' else AMBER_RISK
        ws1.cell(row, 2).value = verdict.upper()
        ws1.cell(row, 2).font  = Font(bold=True, color=v_color, name='Calibri', size=11)
        row += 1
    if trade_win:
        ws1.cell(row, 1).value = 'TRADE WINDOW'; ws1.cell(row, 1).font = _label_font()
        ws1.cell(row, 2).value = trade_win;       ws1.cell(row, 2).font = _value_font()

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2 — Risk Analysis
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Risk Analysis')
    _col_widths(ws2, [14, 42, 62])
    _sheet_header(ws2, 'RISK ANALYSIS — Key Risks & Gap Analysis', 3, row=1)

    _section_label(ws2, 3, 1, 'KEY RISKS')
    ws2.merge_cells('A3:C3')

    for col, hdr in enumerate(['SEVERITY', 'TITLE', 'DESCRIPTION'], 1):
        ws2.cell(4, col).value = hdr
        ws2.cell(4, col).font  = _label_font()
        ws2.cell(4, col).fill  = _fill(BORDER_HEX)
        ws2.cell(4, col).border = _border()

    for i, risk in enumerate(key_risks, 5):
        sev   = risk.get('severity', 'medium')
        title = risk.get('title', '')
        desc  = risk.get('description', '')

        ws2.cell(i, 1).value = sev.upper(); ws2.cell(i, 1).font = _severity_font(sev)
        ws2.cell(i, 1).fill  = _severity_fill(sev)
        ws2.cell(i, 2).value = title
        ws2.cell(i, 2).font  = Font(bold=True, color=DARK, name='Calibri', size=10)
        ws2.cell(i, 2).fill  = _severity_fill(sev)
        ws2.cell(i, 3).value = desc
        ws2.cell(i, 3).font  = _value_font()
        ws2.cell(i, 3).fill  = _severity_fill(sev)
        ws2.cell(i, 3).alignment = Alignment(wrap_text=True, vertical='top')
        ws2.row_dimensions[i].height = 52
        _apply_border_row(ws2, i, 3)

    gap_start = 5 + len(key_risks) + 2
    _section_label(ws2, gap_start, 1, 'GAP ANALYSIS')
    ws2.merge_cells(f'A{gap_start}:C{gap_start}')

    for j, gap in enumerate(gaps, gap_start + 1):
        ws2.merge_cells(f'A{j}:C{j}')
        ws2.cell(j, 1).value = f'• {gap}'
        ws2.cell(j, 1).font  = _value_font()
        ws2.cell(j, 1).fill  = _fill(LIGHT_BG)
        ws2.cell(j, 1).alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[j].height = 32
        _apply_border_row(ws2, j, 3)

    sup_start = gap_start + len(gaps) + 3
    _section_label(ws2, sup_start, 1, 'SUPPLIER & LOGISTICS PROFILE')
    ws2.merge_cells(f'A{sup_start}:C{sup_start}')

    sup_rows = [
        ('GPS Coverage',      f'{gps_pct}%'),
        ('Origin Port',       origin_port),
        ('Destination Port',  dest_port),
        ('Transit Days',      str(transit)),
    ]
    for k, (lbl, val) in enumerate(sup_rows, sup_start + 1):
        ws2.cell(k, 1).value = lbl; ws2.cell(k, 1).font = _label_font()
        ws2.cell(k, 1).fill  = _fill(LIGHT_BG)
        ws2.cell(k, 2).value = val; ws2.cell(k, 2).font = _value_font()
        ws2.merge_cells(f'B{k}:C{k}')
        _apply_border_row(ws2, k, 3)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3 — Recommended Actions
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Recommended Actions')
    _col_widths(ws3, [14, 20, 80])
    _sheet_header(ws3, 'RECOMMENDED ACTIONS', 3, row=1)

    for col, hdr in enumerate(['PRIORITY', 'TIMELINE', 'ACTION'], 1):
        ws3.cell(3, col).value = hdr
        ws3.cell(3, col).font  = _label_font()
        ws3.cell(3, col).fill  = _fill(BORDER_HEX)
        ws3.cell(3, col).border = _border()

    for i, act in enumerate(actions, 4):
        priority = act.get('priority', 'medium')
        timeline = act.get('timeline', '')
        action   = act.get('action', '')
        row_fill = _severity_fill(priority)

        ws3.cell(i, 1).value = priority.upper(); ws3.cell(i, 1).font = _severity_font(priority)
        ws3.cell(i, 1).fill  = row_fill

        ws3.cell(i, 2).value = timeline
        ws3.cell(i, 2).font  = Font(bold=True, color=WHITE, name='Courier New', size=9)
        ws3.cell(i, 2).fill  = _fill(AMBER)
        ws3.cell(i, 2).alignment = Alignment(horizontal='center', vertical='center')

        ws3.cell(i, 3).value = action
        ws3.cell(i, 3).font  = _value_font()
        ws3.cell(i, 3).fill  = row_fill
        ws3.cell(i, 3).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.row_dimensions[i].height = 42
        _apply_border_row(ws3, i, 3)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
